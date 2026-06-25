from flask import Blueprint, jsonify, Response, request
from .auth import require_auth
from .models import MetaModelDefinition, MetaModelRelease, MetaModelReleaseItem, Domain
import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO

bp_export = Blueprint('export', __name__, url_prefix='/api/export')
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")


def _get_release_models(release_no: str):
    rel = MetaModelRelease.query.filter_by(release_no=release_no).first()
    if not rel:
        return None, None
    items = MetaModelReleaseItem.query.filter_by(release_id=rel.id).all()
    ids = [i.model_definition_id for i in items]
    models = MetaModelDefinition.query.filter(MetaModelDefinition.id.in_(ids)).all()
    result = {
        'version': release_no,
        'exported_at': rel.released_at.isoformat() if rel.released_at else None,
    }
    for m in models:
        result[m.model_type] = m.content_json
    return result, rel


@bp_export.get('/json/<release_no>')
@require_auth()
def export_json(release_no):
    result, rel = _get_release_models(release_no)
    if result is None:
        return jsonify({'error': 'not found'}), 404
    return jsonify(result)


@bp_export.get('/yaml/<release_no>')
@require_auth()
def export_yaml(release_no):
    result, rel = _get_release_models(release_no)
    if result is None:
        return jsonify({'error': 'not found'}), 404
    yaml_str = yaml.dump(result, allow_unicode=True, default_flow_style=False, sort_keys=False)
    return Response(yaml_str, mimetype='text/yaml')


def _write_sheet(ws, headers: list[str], rows: list[list]):
    """写入表头行+蓝色样式，然后写入数据行。"""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=str(val) if val is not None else '')


def _json_str(val):
    """将值转为 JSON 字符串（用于 语义(JSON) / 步骤(JSON) 列）。"""
    if val is None:
        return ''
    import json
    if isinstance(val, str):
        return val
    return json.dumps(val, ensure_ascii=False)


def _build_xlsx(result: dict) -> BytesIO:
    """
    构建与前端 excel-schema.ts 一致的 Excel 工作簿。
    生成 Sheet：A, B, C, EPC, E1-E8（+隐藏引用表 _要素引用表）。
    """
    import json

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 构建引用表数据 ──
    ref_entries = []  # [{id, name, dimension}]

    # ── Sheet A：价值域 ──
    domains = Domain.query.all()
    a_rows = []
    for d in domains:
        ref_entries.append({'id': d.name, 'name': d.name, 'dimension': 'A'})
        a_rows.append([
            d.name,
            d.name,
            '',
            d.description or '',
            _json_str(d.tags) if hasattr(d, 'tags') else '',
        ])
    ws_a = wb.create_sheet('A')
    _write_sheet(ws_a, ['ID', '名称', '英文名', '描述', '语义(JSON)'], a_rows)

    # ── Sheet B：能力 ──
    b = result.get('behavioral', {})
    actions = b.get('actions', [])
    b_rows = []
    for act in actions:
        aid = act.get('id', '')
        aname = act.get('name', '')
        ref_entries.append({'id': aid, 'name': aname, 'dimension': 'B'})
        b_rows.append([
            aid,
            aname,
            '',
            act.get('description', '') or act.get('output', ''),
            _json_str(act.get('input', '')),
            act.get('domain', ''),
        ])
    ws_b = wb.create_sheet('B')
    _write_sheet(ws_b, ['ID', '名称', '英文名', '描述', '语义(JSON)', '父节点ID'], b_rows)

    # ── Sheet C：场景 ──
    ws_c = wb.create_sheet('C')
    _write_sheet(ws_c, ['ID', '名称', '英文名', '描述', '语义(JSON)', '父节点ID'], [])

    # ── Sheet EPC ──
    epc = result.get('epc', {})
    steps = epc.get('steps', [])
    epc_rows = []
    for step in steps:
        sid = step.get('id', step.get('event_trigger', ''))
        sname = step.get('action', '')
        ref_entries.append({'id': sid, 'name': sname, 'dimension': 'EPC'})
        epc_rows.append([
            sid,
            sname,
            '',
            '',
            '',
            '',
            '',
            _json_str([step]),
        ])
    ws_epc = wb.create_sheet('EPC')
    _write_sheet(ws_epc, [
        'ID', '名称', '英文名', '描述', '语义(JSON)', '父节点ID', '归属场景ID', '步骤(JSON)',
    ], epc_rows)

    # ── Sheet E1：数据（来自 structural.entities） ──
    s = result.get('structural', {})
    entities = s.get('entities', [])
    e1_rows = []
    for e in entities:
        eid = e.get('id', '')
        ename = e.get('name', '')
        ref_entries.append({'id': eid, 'name': ename, 'dimension': 'E1'})
        e1_rows.append([
            eid,
            ename,
            '',
            'E1',
            'project',
            e.get('description', ''),
        ])
    ws_e1 = wb.create_sheet('E1')
    _write_sheet(ws_e1, ['ID', '名称', '英文名', '维度', '可见性', '描述'], e1_rows)

    # ── Sheet E2：行为 ──
    state_machines = b.get('stateMachines', [])
    e2_rows = []
    for sm in state_machines:
        smid = sm.get('id', '')
        smname = sm.get('name', '')
        ref_entries.append({'id': smid, 'name': smname, 'dimension': 'E2'})
        e2_rows.append([
            smid,
            smname,
            '',
            'E2',
            'project',
            '',
        ])
    ws_e2 = wb.create_sheet('E2')
    _write_sheet(ws_e2, ['ID', '名称', '英文名', '维度', '可见性', '描述'], e2_rows)

    # ── Sheet E3：规则 ──
    rules = result.get('rules', {})
    all_rules = []
    for v in rules.get('validations', []):
        all_rules.append(v)
    for g in rules.get('guardrails', []):
        all_rules.append(g)
    for p in rules.get('policies', []):
        all_rules.append(p)
    for pr in rules.get('probes', []):
        all_rules.append(pr)

    e3_rows = []
    for rule in all_rules:
        rid = rule.get('id', '')
        rname = rule.get('name', '')
        ref_entries.append({'id': rid, 'name': rname, 'dimension': 'E3'})
        e3_rows.append([
            rid,
            rname,
            '',
            'E3',
            'project',
            rule.get('expression', rule.get('condition', '')),
        ])
    ws_e3 = wb.create_sheet('E3')
    _write_sheet(ws_e3, ['ID', '名称', '英文名', '维度', '可见性', '描述'], e3_rows)

    # ── Sheet E4：事件 ──
    events = result.get('events', {})
    event_types = events.get('eventTypes', [])
    e4_rows = []
    for et in event_types:
        etid = et.get('id', '')
        etname = et.get('name', '')
        ref_entries.append({'id': etid, 'name': etname, 'dimension': 'E4'})
        e4_rows.append([
            etid,
            etname,
            '',
            'E4',
            'project',
            et.get('severity', ''),
        ])
    ws_e4 = wb.create_sheet('E4')
    _write_sheet(ws_e4, ['ID', '名称', '英文名', '维度', '可见性', '描述'], e4_rows)

    # ── Sheet E5：岗位角色 ──
    ws_e5 = wb.create_sheet('E5')
    _write_sheet(ws_e5, ['ID', '名称', '英文名', '维度', '可见性', '描述'], [])

    # ── Sheet E6：指标（来自 behavioral.indicators） ──
    indicators = b.get('indicators', [])
    e6_rows = []
    for ind in indicators:
        iid = ind.get('id', '')
        iname = ind.get('name', '')
        ref_entries.append({'id': iid, 'name': iname, 'dimension': 'E6'})
        e6_rows.append([
            iid,
            iname,
            '',
            'E6',
            'project',
            ind.get('formula', ''),
        ])
    ws_e6 = wb.create_sheet('E6')
    _write_sheet(ws_e6, ['ID', '名称', '英文名', '维度', '可见性', '描述'], e6_rows)

    # ── Sheet E7：边界约束 ──
    ws_e7 = wb.create_sheet('E7')
    _write_sheet(ws_e7, ['ID', '名称', '英文名', '维度', '可见性', '描述'], [])

    # ── Sheet E8：数据源（来自 interfaces.apis/queries/compute） ──
    interfaces = result.get('interfaces', {})
    all_ifaces = []
    for api in interfaces.get('apis', []):
        all_ifaces.append(api)
    for q in interfaces.get('queries', []):
        all_ifaces.append(q)

    e8_rows = []
    for iface in all_ifaces:
        fid = iface.get('id', '')
        fname = iface.get('name', '')
        ref_entries.append({'id': fid, 'name': fname, 'dimension': 'E8'})
        e8_rows.append([
            fid,
            fname,
            '',
            'E8',
            'project',
            iface.get('url', iface.get('template', '')),
        ])
    ws_e8 = wb.create_sheet('E8')
    _write_sheet(ws_e8, ['ID', '名称', '英文名', '维度', '可见性', '描述'], e8_rows)

    # ── 隐藏引用表 _要素引用表 ──
    ws_ref = wb.create_sheet('_要素引用表')
    _write_sheet(ws_ref, ['ID', '名称', '维度'],
                 [[e['id'], e['name'], e['dimension']] for e in ref_entries])
    # 隐藏引用表
    ws_ref.sheet_state = 'hidden'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@bp_export.get('/xlsx/<release_no>')
@require_auth()
def export_xlsx(release_no):
    result, rel = _get_release_models(release_no)
    if result is None:
        return jsonify({'error': 'not found'}), 404
    output = _build_xlsx(result)
    return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp_export.route('/xlsx-from-manifest', methods=['GET', 'POST'])
@require_auth()
def export_xlsx_from_manifest():
    """直接从 manifest JSON 生成 Excel（无需 release_no）。"""
    payload = request.get_json(force=True)
    result = {
        'version': payload.get('version', payload.get('metadata', {}).get('version', '0.0.0')),
        'exported_at': None,
        'structural': payload.get('structural', payload.get('data', {})),
        'behavioral': payload.get('behavioral', {}),
        'rules': payload.get('rules', {}),
        'events': payload.get('events', {}),
        'interfaces': payload.get('interfaces', {}),
        'epc': payload.get('epc', {}),
    }
    output = _build_xlsx(result)
    return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
